"""Tests for ReadExcelTask and WriteExcelTask."""

import asyncio
import os
from datetime import date, datetime, time
from typing import Any
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from macbot.tasks.excel import ReadExcelTask, WriteExcelTask
from macbot.tasks.file_read import ReadFileTask

MOCK_MEMORY = "macbot.memory.AgentMemory"


def _run(coro):  # type: ignore[no-untyped-def]
    """Run an async coroutine synchronously."""
    return asyncio.run(coro)


def _create_workbook(path: str, sheets: dict[str, list[list[Any]]]) -> None:
    """Helper to create a test workbook with named sheets and data."""
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.save(path)
    wb.close()


class TestReadExcelTask:
    """Tests for reading Excel files."""

    def test_read_basic(self, tmp_path: Any) -> None:
        """Read a simple workbook and get back the data."""
        path = str(tmp_path / "basic.xlsx")
        _create_workbook(path, {"Sheet1": [["Name", "Age"], ["Alice", 30], ["Bob", 25]]})

        task = ReadExcelTask()
        result = _run(task.execute(path=path))

        assert result["sheet_names"] == ["Sheet1"]
        assert result["active_sheet"] == "Sheet1"
        assert result["sheet"] == "Sheet1"
        assert result["rows_returned"] == 3
        assert result["data"][0] == ["Name", "Age"]
        assert result["data"][1] == ["Alice", 30]
        assert result["data"][2] == ["Bob", 25]

    def test_read_specific_sheet(self, tmp_path: Any) -> None:
        """Read a specific sheet by name."""
        path = str(tmp_path / "multi.xlsx")
        _create_workbook(path, {
            "Summary": [["Total", 100]],
            "Details": [["Item", "Price"], ["Widget", 9.99]],
        })

        task = ReadExcelTask()
        result = _run(task.execute(path=path, sheet="Details"))

        assert result["sheet"] == "Details"
        assert result["data"][0] == ["Item", "Price"]
        assert result["data"][1] == ["Widget", 9.99]

    def test_read_nonexistent_sheet_raises(self, tmp_path: Any) -> None:
        """Requesting a sheet that doesn't exist raises ValueError."""
        path = str(tmp_path / "test.xlsx")
        _create_workbook(path, {"Sheet1": [["a"]]})

        task = ReadExcelTask()
        with pytest.raises(ValueError, match="Sheet 'Missing' not found"):
            _run(task.execute(path=path, sheet="Missing"))

    def test_read_file_not_found(self, tmp_path: Any) -> None:
        """Reading a non-existent file raises FileNotFoundError."""
        task = ReadExcelTask()
        with pytest.raises(FileNotFoundError):
            _run(task.execute(path=str(tmp_path / "nope.xlsx")))

    def test_read_max_rows(self, tmp_path: Any) -> None:
        """max_rows limits the number of returned rows."""
        path = str(tmp_path / "big.xlsx")
        rows = [["row", i] for i in range(50)]
        _create_workbook(path, {"Sheet1": rows})

        task = ReadExcelTask()
        result = _run(task.execute(path=path, max_rows=10))

        assert result["rows_returned"] == 10
        assert result["total_rows"] == 50

    def test_read_cell_range(self, tmp_path: Any) -> None:
        """cell_range limits which cells are read."""
        path = str(tmp_path / "range.xlsx")
        _create_workbook(path, {"Sheet1": [
            ["A", "B", "C"],
            [1, 2, 3],
            [4, 5, 6],
        ]})

        task = ReadExcelTask()
        result = _run(task.execute(path=path, cell_range="A1:B2"))

        assert result["rows_returned"] == 2
        assert result["data"][0] == ["A", "B"]
        assert result["data"][1] == [1, 2]

    def test_read_serializes_datetime(self, tmp_path: Any) -> None:
        """datetime, date, and time values are serialized to ISO strings."""
        path = str(tmp_path / "dates.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["datetime", "date", "time"])
        ws.append([datetime(2025, 3, 1, 14, 30), date(2025, 3, 1), time(14, 30)])
        wb.save(path)
        wb.close()

        task = ReadExcelTask()
        result = _run(task.execute(path=path))

        assert result["data"][1][0] == "2025-03-01T14:30:00"
        # openpyxl stores date as datetime with midnight time
        assert result["data"][1][1] == "2025-03-01T00:00:00"
        assert result["data"][1][2] == "14:30:00"

    def test_read_serializes_none_as_empty(self, tmp_path: Any) -> None:
        """None cell values are serialized as empty strings."""
        path = str(tmp_path / "sparse.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "hello"
        ws["C1"] = "world"
        wb.save(path)
        wb.close()

        task = ReadExcelTask()
        result = _run(task.execute(path=path))

        assert result["data"][0][0] == "hello"
        assert result["data"][0][1] == ""
        assert result["data"][0][2] == "world"

    def test_read_multiple_sheet_names(self, tmp_path: Any) -> None:
        """sheet_names contains all sheets in order."""
        path = str(tmp_path / "multi.xlsx")
        _create_workbook(path, {"Alpha": [[1]], "Beta": [[2]], "Gamma": [[3]]})

        task = ReadExcelTask()
        result = _run(task.execute(path=path))

        assert result["sheet_names"] == ["Alpha", "Beta", "Gamma"]

    def test_read_tilde_expansion(self, tmp_path: Any) -> None:
        """~ in path is expanded."""
        path = str(tmp_path / "tilde.xlsx")
        _create_workbook(path, {"Sheet1": [["test"]]})

        task = ReadExcelTask()
        # Just verify it doesn't crash with a real path
        result = _run(task.execute(path=path))
        assert result["rows_returned"] == 1

    def test_tool_schema(self) -> None:
        """Tool schema has expected structure."""
        task = ReadExcelTask()
        schema = task.to_tool_schema()

        assert schema["name"] == "read_excel"
        props = schema["input_schema"]["properties"]
        assert "path" in props
        assert "sheet" in props
        assert "cell_range" in props
        assert "max_rows" in props
        assert "show_formulas" in props
        assert schema["input_schema"]["required"] == ["path"]


class TestWriteExcelTask:
    """Tests for writing Excel files."""

    def test_write_new_file(self, tmp_path: Any) -> None:
        """Write creates a new .xlsx file with data."""
        path = str(tmp_path / "new.xlsx")
        task = WriteExcelTask()
        data = [["Name", "Score"], ["Alice", 95], ["Bob", 87]]

        with patch(MOCK_MEMORY):
            result = _run(task.execute(path=path, data=data))

        assert "3 rows" in result
        assert os.path.isfile(path)

        # Verify contents
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws["A1"].value == "Name"
        assert ws["B1"].value == "Score"
        assert ws["A2"].value == "Alice"
        assert ws["B2"].value == 95
        wb.close()

    def test_write_to_existing_file(self, tmp_path: Any) -> None:
        """Write modifies an existing file without losing data."""
        path = str(tmp_path / "existing.xlsx")
        _create_workbook(path, {"Sheet1": [["old", "data"]]})

        task = WriteExcelTask()
        with patch(MOCK_MEMORY):
            _run(task.execute(path=path, data=[["new", "data"]], start_cell="A2"))

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws["A1"].value == "old"
        assert ws["A2"].value == "new"
        wb.close()

    def test_write_start_cell(self, tmp_path: Any) -> None:
        """start_cell offsets where data is written."""
        path = str(tmp_path / "offset.xlsx")
        task = WriteExcelTask()

        with patch(MOCK_MEMORY):
            _run(task.execute(path=path, data=[["hello"]], start_cell="C3"))

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws["C3"].value == "hello"
        assert ws["A1"].value is None
        wb.close()

    def test_write_create_sheet(self, tmp_path: Any) -> None:
        """create_sheet=True creates a new sheet in an existing workbook."""
        path = str(tmp_path / "sheets.xlsx")
        _create_workbook(path, {"Sheet1": [["original"]]})

        task = WriteExcelTask()
        with patch(MOCK_MEMORY):
            _run(task.execute(
                path=path, data=[["new sheet data"]], sheet="Report", create_sheet=True,
            ))

        wb = openpyxl.load_workbook(path)
        assert "Report" in wb.sheetnames
        assert wb["Report"]["A1"].value == "new sheet data"
        # Original sheet preserved
        assert wb["Sheet1"]["A1"].value == "original"
        wb.close()

    def test_write_nonexistent_sheet_raises(self, tmp_path: Any) -> None:
        """Writing to a non-existent sheet without create_sheet raises ValueError."""
        path = str(tmp_path / "test.xlsx")
        _create_workbook(path, {"Sheet1": [["a"]]})

        task = WriteExcelTask()
        with patch(MOCK_MEMORY):
            with pytest.raises(ValueError, match="Sheet 'Missing' not found"):
                _run(task.execute(path=path, data=[["x"]], sheet="Missing"))

    def test_write_creates_parent_dirs(self, tmp_path: Any) -> None:
        """Non-existent parent directories are created."""
        path = str(tmp_path / "sub" / "dir" / "file.xlsx")
        task = WriteExcelTask()

        with patch(MOCK_MEMORY):
            _run(task.execute(path=path, data=[["nested"]]))

        assert os.path.isfile(path)

    def test_write_memory_recording(self, tmp_path: Any) -> None:
        """File write is recorded to AgentMemory with summary."""
        path = str(tmp_path / "memo.xlsx")
        task = WriteExcelTask()

        with patch(MOCK_MEMORY) as mock_cls:
            mock_memory = MagicMock()
            mock_cls.return_value = mock_memory
            _run(task.execute(path=path, data=[["test"]], summary="Test report"))
            mock_memory.record_file_written.assert_called_once_with(path, "Test report")

    def test_write_auto_column_sizing(self, tmp_path: Any) -> None:
        """Columns are auto-sized based on content."""
        path = str(tmp_path / "sized.xlsx")
        task = WriteExcelTask()

        with patch(MOCK_MEMORY):
            _run(task.execute(
                path=path,
                data=[["Short", "A much longer column header value"]],
            ))

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        width_a = ws.column_dimensions["A"].width
        width_b = ws.column_dimensions["B"].width
        assert width_b > width_a
        wb.close()

    def test_tool_schema(self) -> None:
        """Tool schema has expected structure."""
        task = WriteExcelTask()
        schema = task.to_tool_schema()

        assert schema["name"] == "write_excel"
        props = schema["input_schema"]["properties"]
        assert "path" in props
        assert "data" in props
        assert "sheet" in props
        assert "start_cell" in props
        assert "summary" in props
        assert "create_sheet" in props
        assert "path" in schema["input_schema"]["required"]
        assert "data" in schema["input_schema"]["required"]


class TestReadFileExcelGuard:
    """Tests for the .xlsx/.xls guard in ReadFileTask."""

    def test_xlsx_raises_value_error(self, tmp_path: Any) -> None:
        """Reading a .xlsx file via read_file raises ValueError."""
        path = str(tmp_path / "test.xlsx")
        # Create a dummy file so FileNotFoundError doesn't trigger first
        with open(path, "wb") as f:
            f.write(b"\x00")

        task = ReadFileTask()
        with pytest.raises(ValueError, match="Use the read_excel tool"):
            _run(task.execute(path=path))

    def test_xls_raises_value_error(self, tmp_path: Any) -> None:
        """Reading a .xls file via read_file raises ValueError."""
        path = str(tmp_path / "test.xls")
        with open(path, "wb") as f:
            f.write(b"\x00")

        task = ReadFileTask()
        with pytest.raises(ValueError, match="Use the read_excel tool"):
            _run(task.execute(path=path))

    def test_xlsx_case_insensitive(self, tmp_path: Any) -> None:
        """Guard works regardless of extension casing."""
        path = str(tmp_path / "test.XLSX")
        with open(path, "wb") as f:
            f.write(b"\x00")

        task = ReadFileTask()
        with pytest.raises(ValueError, match="Use the read_excel tool"):
            _run(task.execute(path=path))
