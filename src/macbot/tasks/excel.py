"""Task: Read and Write Excel files

Read spreadsheet data from .xlsx files and create/modify Excel workbooks using openpyxl.
"""

import logging
import os
from datetime import date, datetime, time
from typing import Any

from macbot.tasks.base import Task
from macbot.tasks.registry import task_registry

logger = logging.getLogger(__name__)


def _serialize_cell(value: Any) -> Any:
    """Convert cell value to a JSON-safe type."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return value


def _parse_cell_ref(cell_ref: str) -> tuple[int, int]:
    """Parse a cell reference like 'A1' into (row, col) 1-indexed."""
    col_str = ""
    row_str = ""
    for ch in cell_ref:
        if ch.isalpha():
            col_str += ch.upper()
        else:
            row_str += ch
    col = 0
    for ch in col_str:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return int(row_str), col


class ReadExcelTask(Task):
    """Read data from an Excel (.xlsx) file.

    Example:
        task = ReadExcelTask()
        result = await task.execute(path="~/Documents/report.xlsx")
        # Returns dict with sheet_names, data rows, dimensions, etc.
    """

    @property
    def name(self) -> str:
        return "read_excel"

    @property
    def description(self) -> str:
        return (
            "Read data from an Excel (.xlsx) file. Returns sheet names, dimensions, "
            "and row data as a list of lists. Specify sheet name and cell range to "
            "read a subset. Use show_formulas=True to see formulas instead of values."
        )

    async def execute(
        self,
        path: str,
        sheet: str = "",
        cell_range: str = "",
        max_rows: int = 100,
        show_formulas: bool = False,
    ) -> dict[str, Any]:
        """Read data from an Excel file.

        Args:
            path: Path to the .xlsx file.
            sheet: Sheet name to read (empty = active sheet).
            cell_range: Cell range like "A1:D10" (empty = all data).
            max_rows: Maximum rows to return (default 100).
            show_formulas: If True, show formulas instead of cached values.

        Returns:
            Dict with sheet_names, active_sheet, sheet, dimensions,
            total_rows, rows_returned, and data (list of row lists).

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If the sheet name is not found.
        """
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError(
                "openpyxl is required for Excel support. Install it with: pip install openpyxl"
            )

        path = os.path.expanduser(path)
        if not os.path.exists(path):
            raise FileNotFoundError(f"File not found: {path}")

        wb = openpyxl.load_workbook(path, data_only=not show_formulas)
        try:
            sheet_names = wb.sheetnames
            active_sheet_name = wb.active.title if wb.active else sheet_names[0]

            target_name = sheet if sheet else active_sheet_name
            if target_name not in sheet_names:
                raise ValueError(
                    f"Sheet '{target_name}' not found. Available: {sheet_names}"
                )

            ws = wb[target_name]
            dimensions = ws.dimensions or ""

            if cell_range:
                rows_iter = ws[cell_range]
            else:
                rows_iter = ws.iter_rows()

            data: list[list[Any]] = []
            total_rows = 0
            for row in rows_iter:
                total_rows += 1
                if total_rows <= max_rows:
                    data.append([_serialize_cell(cell.value) for cell in row])

            return {
                "sheet_names": sheet_names,
                "active_sheet": active_sheet_name,
                "sheet": target_name,
                "dimensions": dimensions,
                "total_rows": total_rows,
                "rows_returned": len(data),
                "data": data,
            }
        finally:
            wb.close()


class WriteExcelTask(Task):
    """Write data to an Excel (.xlsx) file.

    Example:
        task = WriteExcelTask()
        result = await task.execute(
            path="~/Documents/report.xlsx",
            data=[["Name", "Score"], ["Alice", 95], ["Bob", 87]],
            summary="Student scores report",
        )
    """

    @property
    def name(self) -> str:
        return "write_excel"

    @property
    def description(self) -> str:
        return (
            "Write data to an Excel (.xlsx) file. Creates a new file or modifies an existing one. "
            "Data is a list of rows (each row a list of cell values). "
            "Supports writing to a specific sheet and start cell. "
            "Include a summary describing the document purpose."
        )

    async def execute(
        self,
        path: str,
        data: list[list[Any]],
        sheet: str = "",
        start_cell: str = "A1",
        summary: str = "",
        create_sheet: bool = False,
    ) -> str:
        """Write data to an Excel file.

        Args:
            path: Path to the .xlsx file (created if it doesn't exist).
            data: List of rows, each row a list of cell values.
            sheet: Target sheet name (empty = active sheet or "Sheet1" for new files).
            start_cell: Top-left cell to start writing at (default "A1").
            summary: Brief description of the document (e.g., "Monthly sales report").
            create_sheet: If True, create the sheet if it doesn't exist.

        Returns:
            Success message with row count and file path.

        Raises:
            ValueError: If the sheet doesn't exist and create_sheet is False.
        """
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError(
                "openpyxl is required for Excel support. Install it with: pip install openpyxl"
            )

        path = os.path.expanduser(path)
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()

        try:
            if sheet:
                if sheet in wb.sheetnames:
                    ws = wb[sheet]
                elif create_sheet:
                    ws = wb.create_sheet(title=sheet)
                else:
                    raise ValueError(
                        f"Sheet '{sheet}' not found. Available: {wb.sheetnames}. "
                        "Set create_sheet=True to create it."
                    )
            else:
                ws = wb.active or wb.worksheets[0]

            start_row, start_col = _parse_cell_ref(start_cell)

            for r_idx, row_data in enumerate(data):
                for c_idx, value in enumerate(row_data):
                    ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)

            # Auto-size columns based on content width
            for col_cells in ws.columns:
                max_length = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            wb.save(path)
        finally:
            wb.close()

        from macbot.memory import AgentMemory

        memory = AgentMemory()
        memory.record_file_written(path, summary)

        return f"Successfully wrote {len(data)} rows to {path}"


# Auto-register on import
task_registry.register(ReadExcelTask())
task_registry.register(WriteExcelTask())
